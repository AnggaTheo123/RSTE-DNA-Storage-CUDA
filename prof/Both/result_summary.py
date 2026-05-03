#!/usr/bin/env python3
"""
result_summary.py
Generate result_summary.xlsx from prof/ CSVs.

Usage:
    python docs/result_summary.py

Outputs:
    docs/result_summary.xlsx (per-kernel time, total time, memory, speedup table)
    docs/chart_runtime.png   (line chart: SEQ vs CUDA across datasets)
    docs/chart_kernel.png    (bar chart: per-kernel breakdown)

Requires: pandas, openpyxl, matplotlib
"""

from __future__ import annotations
import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
PROF_SEQ = ROOT / "prof" / "Sequential"
PROF_PAR = ROOT / "prof" / "Parallel"
PROF_BOTH = ROOT / "prof" / "Both"


def read_csv_rows(p: Path) -> list[dict]:
    if not p.exists():
        return []
    with p.open(newline="") as f:
        return list(csv.DictReader(f))


def collect():
    seq_rows = []
    for f in sorted(PROF_SEQ.glob("*.csv")):
        seq_rows.extend(read_csv_rows(f))

    par_rows = []
    for f in sorted(PROF_PAR.glob("*.csv")):
        par_rows.extend(read_csv_rows(f))

    return seq_rows, par_rows


def main():
    seq, par = collect()
    print(f"[seq] rows: {len(seq)}")
    print(f"[par] rows: {len(par)}")

    if not seq and not par:
        print("No CSV data found in prof/. Run the experiments first.")
        sys.exit(1)

    try:
        import pandas as pd
    except ImportError:
        print("pandas not installed; writing plain CSV summary instead.")
        out = PROF_BOTH / "result_summary.csv"
        with out.open("w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["impl", "dataset", "file_name", "size_bytes",
                        "total_ms", "encoding_rate_bits_per_nt",
                        "rll_max", "rll_violations", "dna_hash_payload"])
            for r in seq:
                w.writerow(["SEQ", r.get("dataset"), r.get("file_name"),
                            r.get("file_size_bytes"), r.get("total_ms"),
                            r.get("encoding_rate_bits_per_nt"),
                            r.get("rll_max"), r.get("rll_violations"),
                            r.get("dna_hash_payload")])
            for r in par:
                w.writerow(["CUDA", r.get("dataset"), r.get("file_name"),
                            r.get("file_size_bytes"), r.get("total_ms"),
                            r.get("encoding_rate_bits_per_nt"),
                            r.get("rll_max"), r.get("rll_violations"),
                            r.get("dna_hash_payload")])
        print(f"Wrote {out}")
        return

    df_seq = pd.DataFrame(seq)
    df_par = pd.DataFrame(par)

    if not df_seq.empty:
        df_seq["impl"] = "SEQ"
    if not df_par.empty:
        df_par["impl"] = "CUDA"

    out_xlsx = PROF_BOTH / "result_summary.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xw:
        if not df_seq.empty:
            df_seq.to_excel(xw, sheet_name="Sequential", index=False)
        if not df_par.empty:
            df_par.to_excel(xw, sheet_name="Parallel", index=False)

        if not df_seq.empty and not df_par.empty:
            on = ["dataset", "file_name"]
            join = pd.merge(
                df_seq[on + ["total_ms", "dna_hash_payload"]].rename(
                    columns={"total_ms": "seq_total_ms",
                             "dna_hash_payload": "seq_hash"}
                ),
                df_par[on + ["total_ms", "dna_hash_payload"]].rename(
                    columns={"total_ms": "cuda_total_ms",
                             "dna_hash_payload": "cuda_hash"}
                ),
                on=on,
                how="inner",
            )
            join["seq_total_ms"]  = pd.to_numeric(join["seq_total_ms"],  errors="coerce")
            join["cuda_total_ms"] = pd.to_numeric(join["cuda_total_ms"], errors="coerce")
            join["speedup"]    = join["seq_total_ms"] / join["cuda_total_ms"]
            join["hash_match"] = join["seq_hash"] == join["cuda_hash"]
            join.to_excel(xw, sheet_name="Summary", index=False)

    print(f"Wrote {out_xlsx}")

    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        print("matplotlib not installed; skipping charts.")
        return

    if df_seq.empty or df_par.empty:
        return

    # Coerce numeric columns
    seq_phase_cols = ["total_ms", "lsbm_ms", "huffman_ms", "encode_ms"]
    par_phase_cols = ["total_ms", "h2d_ms", "bits_kernel_ms", "lsbm_kernel_ms",
                      "huffman_cpu_ms", "encode_kernel_ms", "constraint_kernel_ms",
                      "assembly_ms", "d2h_ms"]
    for c in seq_phase_cols:
        if c in df_seq.columns:
            df_seq[c] = pd.to_numeric(df_seq[c], errors="coerce")
    for c in par_phase_cols:
        if c in df_par.columns:
            df_par[c] = pd.to_numeric(df_par[c], errors="coerce")

    # SEQ phase grouping (paper-side phases) and CUDA phase grouping (kernel-side)
    seq_phase_map = {
        "LSBM (KMP)":  "lsbm_ms",
        "Huffman":     "huffman_ms",
        "Encode+Cstr": "encode_ms",
    }
    cuda_phase_map = {
        "H2D":         "h2d_ms",
        "Bits kern.":  "bits_kernel_ms",
        "LSBM kern.":  "lsbm_kernel_ms",
        "Huffman CPU": "huffman_cpu_ms",
        "Encode kern.":"encode_kernel_ms",
        "Cstr kern.":  "constraint_kernel_ms",
        "Assembly":    "assembly_ms",
        "D2H":         "d2h_ms",
    }

    datasets = sorted(set(df_seq["dataset"]) & set(df_par["dataset"]))

    # ------------------------------------------------------------------
    # 1) Per-dataset PNG: 2 panels side-by-side
    #    Left  = SEQ phase breakdown (bar)
    #    Right = CUDA phase breakdown (bar)
    # ------------------------------------------------------------------
    for d in datasets:
        seq_row  = df_seq [df_seq ["dataset"] == d].mean(numeric_only=True)
        cuda_row = df_par [df_par ["dataset"] == d].mean(numeric_only=True)

        seq_labels = list(seq_phase_map.keys())
        seq_vals   = [seq_row.get(seq_phase_map[k], 0) for k in seq_labels]

        cuda_labels = [k for k in cuda_phase_map
                       if cuda_phase_map[k] in cuda_row.index]
        cuda_vals   = [cuda_row[cuda_phase_map[k]] for k in cuda_labels]

        seq_total  = float(seq_row.get("total_ms", sum(seq_vals)))
        cuda_total = float(cuda_row.get("total_ms", sum(cuda_vals)))
        speedup    = seq_total / cuda_total if cuda_total > 0 else 0

        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        # SEQ panel
        bars = axes[0].bar(seq_labels, seq_vals,
                           color=["#1f77b4", "#9467bd", "#ff7f0e"])
        axes[0].set_title(f"Sequential — total {seq_total:.1f} ms",
                          fontsize=11, fontweight="bold")
        axes[0].set_ylabel("Time (ms)")
        for b, v in zip(bars, seq_vals):
            axes[0].text(b.get_x() + b.get_width() / 2, v,
                         f"{v:.1f}", ha="center", va="bottom", fontsize=9)
        axes[0].tick_params(axis="x", rotation=15)

        # CUDA panel
        cuda_colors = ["#aec7e8", "#1f77b4", "#d62728", "#9467bd",
                       "#ff7f0e", "#2ca02c", "#8c564b", "#7f7f7f"][: len(cuda_labels)]
        bars = axes[1].bar(cuda_labels, cuda_vals, color=cuda_colors)
        axes[1].set_title(f"CUDA — total {cuda_total:.1f} ms (speedup {speedup:.1f}×)",
                          fontsize=11, fontweight="bold")
        axes[1].set_ylabel("Time (ms)")
        for b, v in zip(bars, cuda_vals):
            axes[1].text(b.get_x() + b.get_width() / 2, v,
                         f"{v:.2f}", ha="center", va="bottom", fontsize=9)
        axes[1].tick_params(axis="x", rotation=20)

        fig.suptitle(f"RSTE phase breakdown — dataset: {d}",
                     fontsize=13, fontweight="bold")
        fig.tight_layout(rect=[0, 0, 1, 0.96])
        fname = PROF_BOTH / f"chart_dataset_{d}.png"
        fig.savefig(fname, dpi=120)
        plt.close(fig)
        print(f"Wrote {fname}")

    # ------------------------------------------------------------------
    # 2) Combined: total runtime SEQ vs CUDA (linear + log)
    # ------------------------------------------------------------------
    s_y = [df_seq[df_seq["dataset"] == d]["total_ms"].mean() for d in datasets]
    p_y = [df_par[df_par["dataset"] == d]["total_ms"].mean() for d in datasets]
    x = list(range(len(datasets)))
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    for ax, ylog in zip(axes, [False, True]):
        ax.plot(x, s_y, "-o", label="Sequential", color="#d62728", linewidth=2)
        ax.plot(x, p_y, "-s", label="CUDA",       color="#2ca02c", linewidth=2)
        ax.set_xticks(x)
        ax.set_xticklabels(datasets, rotation=15)
        ax.set_ylabel("Runtime (ms)" + (" — log scale" if ylog else ""))
        if ylog:
            ax.set_yscale("log")
        ax.set_title("Linear scale" if not ylog else "Log scale")
        ax.legend()
        ax.grid(True, alpha=0.3)
    fig.suptitle("Total runtime — Sequential vs CUDA", fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    fig.savefig(PROF_BOTH / "chart_runtime.png", dpi=120)
    plt.close(fig)
    print(f"Wrote {PROF_BOTH/'chart_runtime.png'}")

    # ------------------------------------------------------------------
    # 3) Combined: speedup bar chart
    # ------------------------------------------------------------------
    speedup = [s / p if p > 0 else 0 for s, p in zip(s_y, p_y)]
    fig, ax = plt.subplots(figsize=(8, 5))
    bars = ax.bar(datasets, speedup, color="#2ca02c")
    for b, v in zip(bars, speedup):
        ax.text(b.get_x() + b.get_width() / 2, v,
                f"{v:.1f}×", ha="center", va="bottom",
                fontsize=10, fontweight="bold")
    ax.set_ylabel("Speedup (CUDA vs Sequential)")
    ax.set_title("End-to-end speedup per dataset", fontweight="bold")
    ax.tick_params(axis="x", rotation=15)
    ax.grid(True, axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(PROF_BOTH / "chart_speedup.png", dpi=120)
    plt.close(fig)
    print(f"Wrote {PROF_BOTH/'chart_speedup.png'}")

    # ------------------------------------------------------------------
    # 4) Combined: CUDA per-kernel breakdown stacked across datasets
    # ------------------------------------------------------------------
    kernels = ["bits_kernel_ms", "lsbm_kernel_ms", "huffman_cpu_ms",
               "encode_kernel_ms", "constraint_kernel_ms",
               "assembly_ms", "h2d_ms", "d2h_ms"]
    avail = [k for k in kernels if k in df_par.columns]
    if avail:
        means = df_par.groupby("dataset")[avail].mean().reindex(datasets)
        ax = means.plot(kind="bar", stacked=True, figsize=(10, 6),
                        colormap="tab20")
        ax.set_ylabel("Time (ms)")
        ax.set_title("CUDA per-kernel breakdown by dataset",
                     fontweight="bold")
        ax.legend(title="Phase", bbox_to_anchor=(1.02, 1), loc="upper left")
        ax.figure.tight_layout()
        ax.figure.savefig(PROF_PAR / "chart_kernel.png", dpi=120)
        plt.close(ax.figure)
        print(f"Wrote {PROF_PAR/'chart_kernel.png'}")

    # ------------------------------------------------------------------
    # 5) Embed charts into Excel
    # ------------------------------------------------------------------
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as xlImage
        
        wb = load_workbook(out_xlsx)
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            
            try:
                img1 = xlImage(str(PROF_BOTH / "chart_speedup.png"))
                ws.add_image(img1, "A10")
                
                img2 = xlImage(str(PROF_BOTH / "chart_runtime.png"))
                ws.add_image(img2, "K10")
                
                img3 = xlImage(str(PROF_PAR / "chart_kernel.png"))
                ws.add_image(img3, "A40")
                
                wb.save(out_xlsx)
                print("Successfully embedded charts into result_summary.xlsx")
            except Exception as e:
                print(f"Failed to embed images: {e}")
    except ImportError:
        print("openpyxl or PIL not installed properly; could not embed charts.")


if __name__ == "__main__":
    main()
